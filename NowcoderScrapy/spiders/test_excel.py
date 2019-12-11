import os
import xlrd
import pandas
from openpyxl import load_workbook

data = xlrd.open_workbook("test.xlsx")
table = data.sheets()[0]  # 选定表
nrows = table.nrows  # 获取行号

columns = table.row_values(0)
print(columns)
for row in range(1, nrows):
    alldata = table.row_values(row)  # 循环输出excel表中每一行，即所有数据
    print(alldata)
    toFile = './tofile.xlsx'
    file_is_exist = os.path.exists(toFile)
    df = pandas.DataFrame([alldata], columns=columns)
    writer = pandas.ExcelWriter(toFile, engine='openpyxl')
    if not file_is_exist:
        df.to_excel(
            writer,
            header=True,
            sheet_name="test",
            index=False,  # 列最左侧索引
            columns=columns
        )
    else:
        curBook = load_workbook(toFile)
        row_start = curBook["test"].max_row
        print(row_start)
        writer.book = curBook
        # print(dict((ws.title, ws) for ws in curBook.worksheets))
        writer.sheets = dict((ws.title, ws) for ws in curBook.worksheets)
        for ws in curBook.worksheets:
            print((ws.title, ws))
        print(writer.sheets)
        df.to_excel(
            writer,
            header=False,
            sheet_name="test",
            index=False,
            startrow=row_start,
        )
    writer.save()
